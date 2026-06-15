#!/usr/bin/env python3
"""Import Title Core sheet from Offshore Team Workload Tracker.xlsx into workload.json."""

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl

WORKBOOK = Path.home() / "Desktop" / "Offshore Team Workload Tracker.xlsx"
OUTPUT = Path(__file__).resolve().parent.parent / "public" / "data" / "workload.json"


def parse_int(value) -> int:
    if value is None:
        return 0
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return 0


def stories_from_text(text: str) -> int:
    patterns = [
        r"No of stories\s*(?:\([^)]*\))?\s*:\s*(\d+)",
        r"total\s*=\s*(\d+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return int(match.group(1))
    return 0


def bugs_from_text(text: str) -> int:
    match = re.search(r"Bugs worked\s*:\s*(\d+)", text, re.IGNORECASE)
    return int(match.group(1)) if match else 0


def release_count_from_text(text: str) -> int:
    match = re.search(r"Release\s*:\s*(\d+)", text, re.IGNORECASE)
    return int(match.group(1)) if match else 0


def parse_details(details, releases_column: int) -> tuple[list[dict], int]:
    text = str(details).strip() if details else ""
    releases: list[dict] = []

    if text:
        blocks = re.split(r"Release\d+:", text, flags=re.IGNORECASE)
        if len(blocks) > 1:
            for block in blocks[1:]:
                releases.append(
                    {
                        "stories": stories_from_text(block),
                        "bugs": bugs_from_text(block),
                    }
                )

        if not releases:
            stories = stories_from_text(text)
            bugs = bugs_from_text(text)
            if stories > 0 or bugs > 0:
                releases = [{"stories": stories, "bugs": bugs}]

    final_count = releases_column
    if final_count <= 0:
        final_count = release_count_from_text(text)

    if final_count <= 0:
        releases = [r for r in releases if r["stories"] > 0 or r["bugs"] > 0]
        return releases, len(releases)

    if not releases:
        per_release_match = re.search(
            r"(\d+)\s*tasks?\s*in each release", text, re.IGNORECASE
        )
        if per_release_match:
            each = int(per_release_match.group(1))
            releases = [{"stories": each, "bugs": 0} for _ in range(final_count)]
        else:
            stories = stories_from_text(text)
            bugs = bugs_from_text(text)
            releases = [
                {
                    "stories": stories // final_count if stories and final_count else 0,
                    "bugs": bugs if i == 0 else 0,
                }
                for i in range(final_count)
            ]

    while len(releases) < final_count:
        releases.append({"stories": 0, "bugs": 0})
    releases = releases[:final_count]

    return releases, final_count


def norm_month(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    s = str(value).strip().replace("\n", "")
    months = {
        "jan": "2026-01",
        "feb": "2026-02",
        "mar": "2026-03",
        "apr": "2026-04",
        "may": "2026-05",
        "jun": "2026-06",
        "dec": "2026-12",
    }
    low = s.lower()[:3]
    return months.get(low, s)


def split_names(value) -> list[str]:
    if not value:
        return []
    result = []
    for name in re.split(r"[/,]", str(value)):
        name = re.sub(r"\(.*?\)", "", name).strip()
        name = re.sub(r"\s+", " ", name)
        if name and name.lower() not in ("na", "none"):
            result.append(name)
    return result


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK}")

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Title Core"]

    records = []
    for row in range(2, ws.max_row + 1):
        project = ws.cell(row, 4).value
        if not project:
            continue

        releases_column = parse_int(ws.cell(row, 5).value)
        details = ws.cell(row, 6).value
        parsed, release_count = parse_details(details, releases_column)

        if release_count <= 0:
            continue

        records.append(
            {
                "month": norm_month(ws.cell(row, 1).value),
                "primaryResource": ws.cell(row, 2).value,
                "fleet": ws.cell(row, 3).value,
                "project": str(project).strip(),
                "releases": release_count,
                "releaseVersion": str(ws.cell(row, 7).value)
                if ws.cell(row, 7).value
                else None,
                "testPlanLink": str(ws.cell(row, 8).value)
                if ws.cell(row, 8).value
                else None,
                "backupResource": ws.cell(row, 9).value,
                "primaryResources": split_names(ws.cell(row, 2).value),
                "backupResources": split_names(ws.cell(row, 9).value),
                "releaseBreakdown": parsed,
                "totalStories": sum(x["stories"] for x in parsed),
                "totalBugs": sum(x["bugs"] for x in parsed),
            }
        )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": WORKBOOK.name,
        "sheet": "Title Core",
        "lastUpdated": datetime.now().isoformat(),
        "records": records,
    }
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)

    print(f"Imported {len(records)} records -> {OUTPUT}")


if __name__ == "__main__":
    main()
